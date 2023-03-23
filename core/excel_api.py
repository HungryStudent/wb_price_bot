from io import BytesIO
from tempfile import NamedTemporaryFile
from typing import List

from openpyxl.workbook import Workbook
from openpyxl.worksheet import worksheet
from openpyxl import load_workbook
from openpyxl.worksheet.worksheet import Worksheet

from core import schemas


def create_export_data(account: schemas.AccountOut, articles: List[schemas.ArticleOut]):
    wb = Workbook()
    ws = wb.active
    ws.append(["Артикул", "Мин цена"])
    for article in articles:
        ws.append([article.id, article.min_price])
    wb.save("docs/" + account.name + ".xlsx")


def parse_data(file_dir) -> List[schemas.ArticleEdit]:
    wb: Workbook = load_workbook(file_dir, data_only=True)
    ws: Worksheet = wb.active
    data = []
    for row in ws.values:
        if row[0] == "Артикул":
            continue
        if row[0] is None or row[1] is None:
            continue
        try:
            data.append(schemas.ArticleEdit(id=int(row[0]), min_price=int(row[1])))
        except ValueError:
            continue
    return data
